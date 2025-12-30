"""
Exportar resultados a Excel
Portfolio Data Analyst - Nivel Básico

Este script exporta los resultados del análisis a Excel
con múltiples hojas y formato profesional
"""

import pandas as pd
from sqlalchemy import create_engine
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference

# Configuración
DATABASE_URL = 'postgresql://postgres:123456@localhost:5432/retail_analysis_basico'

def exportar_a_excel():
    """Exporta resultados a Excel con formato profesional"""
    
    print("=" * 70)
    print("EXPORTANDO RESULTADOS A EXCEL")
    print("=" * 70)
    
    try:
        # Crear conexión con encoding UTF-8
        engine = create_engine(
            DATABASE_URL,
            connect_args={
                'client_encoding': 'utf8'
            },
            pool_pre_ping=True
        )
        
        # Crear carpeta excel si no existe
        excel_dir = Path(__file__).parent.parent / 'excel'
        excel_dir.mkdir(exist_ok=True)
        excel_path = excel_dir / 'reporte_ventas_basico.xlsx'
        
        # Consultas para diferentes hojas
        queries = {
            'Resumen': """
                SELECT 
                    COUNT(*) AS total_ventas,
                    SUM(total) AS ingresos_totales,
                    AVG(total) AS ticket_promedio,
                    MIN(total) AS venta_minima,
                    MAX(total) AS venta_maxima
                FROM ventas
            """,
            'Por Region': """
                SELECT 
                    region,
                    COUNT(*) AS num_ventas,
                    SUM(total) AS total_ventas,
                    AVG(total) AS promedio_venta
                FROM ventas
                GROUP BY region
                ORDER BY total_ventas DESC
            """,
            'Por Mes': """
                SELECT 
                    DATE_TRUNC('month', fecha) AS mes,
                    COUNT(*) AS num_ventas,
                    SUM(total) AS total_mes,
                    AVG(total) AS promedio_mes
                FROM ventas
                GROUP BY mes
                ORDER BY mes
            """,
            'Top Productos': """
                SELECT 
                    producto,
                    SUM(cantidad) AS unidades_vendidas,
                    SUM(total) AS ingresos_totales,
                    COUNT(*) AS num_ventas
                FROM ventas
                GROUP BY producto
                ORDER BY ingresos_totales DESC
                LIMIT 10
            """,
            'Por Segmento': """
                SELECT 
                    c.segmento,
                    COUNT(*) AS num_ventas,
                    SUM(v.total) AS total_ventas,
                    AVG(v.total) AS ticket_promedio
                FROM ventas v
                JOIN clientes c ON v.cliente_id = c.cliente_id
                GROUP BY c.segmento
                ORDER BY total_ventas DESC
            """
        }
        
        # Crear Excel con múltiples hojas
        print("\nCreando archivo Excel...")
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            for nombre, query in queries.items():
                df = pd.read_sql(query, engine)
                df.to_excel(writer, sheet_name=nombre, index=False)
                print(f"   [OK] Hoja '{nombre}' creada con {len(df)} filas")
        
        # Formatear Excel
        print("\nAplicando formato...")
        wb = load_workbook(excel_path)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Formato de encabezados
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_path)
        
        print(f"\n[EXITO] Reporte Excel creado: {excel_path}")
        print("=" * 70)
        
        return True
        
    except Exception as e:
        print(f"\n[ERROR] Error al exportar a Excel: {e}")
        return False


if __name__ == "__main__":
    exportar_a_excel()

